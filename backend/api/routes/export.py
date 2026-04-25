from datetime import datetime
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Optional

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from backend.core import get_current_active_user
from backend.db import get_db
from backend.models import CollectionPeriod, ScheduleEntry, User, UserRole
from backend.services import build_schedule_summary, build_schedule_validation

router = APIRouter(prefix="/export", tags=["export"])

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
WARNING_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")


def _build_schedule_string(entry: ScheduleEntry) -> str:
    meta = entry.meta or {}

    if entry.status == "shift":
        start = meta.get("shiftStart", "")
        end = meta.get("shiftEnd", "")
        return f"{start}-{end}" if start and end else ""
    if entry.status == "split":
        s1 = meta.get("splitStart1", "")
        e1 = meta.get("splitEnd1", "")
        s2 = meta.get("splitStart2", "")
        e2 = meta.get("splitEnd2", "")
        return f"{s1}-{e1} | {s2}-{e2}" if s1 and e1 and s2 and e2 else ""
    if entry.status == "dayoff":
        return "dayoff"
    if entry.status == "vacation":
        return "vacation"
    return ""


def _prepare_export_rows(users: list[User], period: CollectionPeriod, db: Session) -> tuple[list[dict], list[str]]:
    rows: list[dict] = []
    all_dates = {
        single_day.isoformat()
        for single_day in (
            period.period_start.fromordinal(period.period_start.toordinal() + offset)
            for offset in range((period.period_end - period.period_start).days + 1)
        )
    }

    for user in sorted(users, key=lambda item: (item.full_name or item.email or "").lower()):
        entries = (
            db.query(ScheduleEntry)
            .filter(
                ScheduleEntry.user_id == user.id,
                ScheduleEntry.period_id == period.id,
            )
            .order_by(ScheduleEntry.day.asc())
            .all()
        )

        schedule_map = {entry.day.isoformat(): _build_schedule_string(entry) for entry in entries}
        summary = build_schedule_summary(entries)
        validation = build_schedule_validation(entries)

        rows.append(
            {
                "user": user,
                "entries": entries,
                "schedule_map": schedule_map,
                "summary": summary,
                "validation": validation,
            }
        )

    return rows, sorted(all_dates)


def _style_header(row) -> None:
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 36)


def _generate_excel_file(rows: list[dict], date_columns: list[str]) -> Path:
    workbook = Workbook()

    schedule_sheet = workbook.active
    schedule_sheet.title = "Schedule"
    schedule_headers = ["Alliance", "Employee", "Total hours", "Vacation days", "Violations"]
    schedule_headers.extend(date_columns)
    schedule_sheet.append(schedule_headers)
    _style_header(schedule_sheet[1])

    for row in rows:
        validation = row["validation"]
        violation_codes = ", ".join(item["code"] for item in validation["violations"])
        values = [
            row["user"].alliance or "",
            row["user"].full_name or row["user"].email or f"User {row['user'].id}",
            row["summary"]["period_total_hours"],
            row["summary"]["vacation_days_count"],
            violation_codes,
        ]
        values.extend(row["schedule_map"].get(date_str, "") for date_str in date_columns)
        schedule_sheet.append(values)

        current_row = schedule_sheet.max_row
        if violation_codes:
            for cell in schedule_sheet[current_row]:
                cell.fill = WARNING_FILL

    _autosize_columns(schedule_sheet)
    schedule_sheet.freeze_panes = "A2"

    summary_sheet = workbook.create_sheet("Summary")
    summary_headers = [
        "Employee",
        "Submitted",
        "Total hours",
        "Vacation days",
        "Max work streak",
        "Weekly hours",
        "Violation count",
        "Violation codes",
    ]
    summary_sheet.append(summary_headers)
    _style_header(summary_sheet[1])

    for row in rows:
        summary = row["summary"]
        validation = row["validation"]
        weekly_hours = ", ".join(
            f"{week_start.isoformat()}: {hours}" for week_start, hours in summary["weekly_hours"].items()
        )
        violation_codes = ", ".join(item["code"] for item in validation["violations"])
        summary_sheet.append(
            [
                row["user"].full_name or row["user"].email or f"User {row['user'].id}",
                "yes" if row["entries"] else "no",
                summary["period_total_hours"],
                summary["vacation_days_count"],
                summary["max_work_streak"],
                weekly_hours,
                len(validation["violations"]),
                violation_codes,
            ]
        )

        current_row = summary_sheet.max_row
        if validation["violations"]:
            for cell in summary_sheet[current_row]:
                cell.fill = WARNING_FILL

    _autosize_columns(summary_sheet)
    summary_sheet.freeze_panes = "A2"

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        file_path = Path(temp_file.name)

    workbook.save(file_path)
    return file_path


def _remove_temp_file(path: str) -> None:
    try:
        Path(path).unlink(missing_ok=True)
    except OSError:
        pass


@router.get("/schedule")
def export_schedule(
    background_tasks: BackgroundTasks,
    period_id: Optional[int] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    if current_user.role != UserRole.MANAGER:
        raise HTTPException(status_code=403, detail="РќРµРґРѕСЃС‚Р°С‚РѕС‡РЅРѕ РїСЂР°РІ")

    if period_id:
        period = db.query(CollectionPeriod).filter(CollectionPeriod.id == period_id).first()
        if not period:
            raise HTTPException(status_code=404, detail="РџРµСЂРёРѕРґ РЅРµ РЅР°Р№РґРµРЅ")
        if period.alliance != current_user.alliance:
            raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє СЌС‚РѕРјСѓ РїРµСЂРёРѕРґСѓ")
    else:
        period = (
            db.query(CollectionPeriod)
            .filter(CollectionPeriod.is_open.is_(True), CollectionPeriod.alliance == current_user.alliance)
            .order_by(CollectionPeriod.created_at.desc())
            .first()
        )
        if not period:
            raise HTTPException(status_code=400, detail="РќРµС‚ Р°РєС‚РёРІРЅРѕРіРѕ РїРµСЂРёРѕРґР° СЃР±РѕСЂР°")

    users = (
        db.query(User)
        .filter(
            User.is_verified.is_(True),
            User.alliance == current_user.alliance,
            User.role == UserRole.USER,
        )
        .all()
    )
    if not users:
        raise HTTPException(status_code=404, detail="РќРµС‚ РґР°РЅРЅС‹С… РґР»СЏ СЌРєСЃРїРѕСЂС‚Р°")

    rows, date_columns = _prepare_export_rows(users, period, db)

    try:
        file_path = _generate_excel_file(rows, date_columns)
        background_tasks.add_task(_remove_temp_file, str(file_path))
        return FileResponse(
            path=file_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"schedule_{period.period_start}_{period.period_end}.xlsx",
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"РћС€РёР±РєР° РіРµРЅРµСЂР°С†РёРё С„Р°Р№Р»Р°: {exc}") from exc
